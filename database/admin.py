from django.contrib import admin, messages
from django.contrib.admin.helpers import ACTION_CHECKBOX_NAME
from lxml import etree
from django.http import HttpResponse, JsonResponse
from django.utils.html import format_html
from openpyxl.reader.excel import load_workbook

from database import models
# Register your models here.
from openpyxl import Workbook

from database.models import List
from utils import requests_pro
from import_export.admin import ImportMixin


@admin.register(models.List)
class ListAdmin(ImportMixin, admin.ModelAdmin):
    list_display = ('sentences', 'labels',)
    list_per_page = 20
    actions = ["export_as_excel", 'request_data']

    # 导出数据
    def export_as_excel(self, request, queryset):
        meta = self.model._meta  # 用于定义文件名, 格式为: app名.模型类名
        field_names = [field.name for field in meta.fields]  # 模型所有字段名
        response = HttpResponse(content_type='application/msexcel')  # 定义响应内容类型
        response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'  # 定义响应数据格式
        wb = Workbook()  # 新建Workbook
        ws = wb.active  # 使用当前活动的Sheet表
        ws.append(field_names)  # 将模型字段名作为标题写入第一行
        for obj in queryset:  # 遍历选择的对象列表
            for field in field_names:
                data = [f'{getattr(obj, field)}' for field in field_names]  # 将模型属性值的文本格式组成列表
            ws.append(data)  # 写入模型属性值
        wb.save(response)  # 将数据存入响应内容
        return response
    # def import_as_excel(self,request,queryset):
    #     excel_file = request.FILES.get('excel_file')
    #     if excel_file:
    #         wb = load_workbook(excel_file)
    #         ws = wb.active
    #         for row in ws.iter_rows(min_row=2, values_only=True):  # 从第二行开始读取数据
    #             sentences, labels = row[0], row[1]
    #             List.objects.create(sentences=sentences, labels=labels)
    #
    #         self.message_user(request, "Excel 文件导入成功！")
    #     else:
    #         self.message_user(request, "请选择要导入的 Excel 文件！", level='ERROR')
    #
    # import_as_excel.short_description = "导入 Excel 文件"
    export_as_excel.short_description = '导出表格'  # 该动作在admin中的显示文字
    export_as_excel.type = 'warning'
    export_as_excel.icon = 'el-icon-download'


@admin.register(models.ReqeustsData)
class PieData(admin.ModelAdmin):
    list_display = ('name', 'author', 'reviewer', 'last_comment_time', 'comment', 'labels', 'href_link', 'collection')
    list_filter = ['name', 'labels', ]

    list_per_page = 20
    actions = ["export_as_excel", 'request_data']

    # 导出数据
    def export_as_excel(self, request, queryset):
        meta = self.model._meta  # 用于定义文件名, 格式为: app名.模型类名
        field_names = [field.name for field in meta.fields]  # 模型所有字段名
        response = HttpResponse(content_type='application/msexcel')  # 定义响应内容类型
        response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'  # 定义响应数据格式
        wb = Workbook()  # 新建Workbook
        ws = wb.active  # 使用当前活动的Sheet表
        ws.append(field_names)  # 将模型字段名作为标题写入第一行
        for obj in queryset:  # 遍历选择的对象列表
            for field in field_names:
                data = [f'{getattr(obj, field)}' for field in field_names]  # 将模型属性值的文本格式组成列表
            ws.append(data)  # 写入模型属性值
        wb.save(response)  # 将数据存入响应内容
        return response

    export_as_excel.short_description = '导出表格'  # 该动作在admin中的显示文字
    export_as_excel.type = 'warning'
    export_as_excel.icon = 'el-icon-download'


@admin.register(models.Douyin)
class Douyin(admin.ModelAdmin):
    list_display = ('username', 'comment', 'vedioname','labels', 'href_link')
    list_per_page = 20
    actions = ["export_as_excel"]

    def export_as_excel(self, request, queryset):
        response = HttpResponse(content_type='application/msexcel')
        return response

    export_as_excel.short_description = '导出表格'
    export_as_excel.type = 'warning'
    export_as_excel.icon = 'el-icon-download'


@admin.register(admin.models.LogEntry)
class LogEntryAdmin(admin.ModelAdmin):
    list_display = ['action_time', 'user', 'content_type', '__str__']
    list_display_links = ['action_time']
    list_filter = ['action_time', 'content_type', 'user']
    list_per_page = 15
    readonly_fields = ['action_time', 'user', 'content_type', 'object_id', 'object_repr', 'action_flag',
                       'change_message']


@admin.register(models.Colect)
class Colect(admin.ModelAdmin):  # 这里需要修改

    list_display = ("co_title", "co_down_link", "href_link")
    list_filter = ["co_title", "co_down_link"]

    list_per_page = 20
    actions = ["export_as_excel"]

    # 封面
    def game_picture(self, models_obj):
        return format_html('<img src="{}" height="50" width="50">', '{}'.format(models_obj.co_down_img))

    game_picture.short_description = '游戏LOGO'

    def export_as_excel(self, request, queryset):
        meta = self.model._meta  # 用于定义文件名, 格式为: app名.模型类名
        field_names = [field.name for field in meta.fields]  # 模型所有字段名
        response = HttpResponse(content_type='application/msexcel')  # 定义响应内容类型
        response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'  # 定义响应数据格式
        wb = Workbook()  # 新建Workbook
        ws = wb.active  # 使用当前活动的Sheet表
        ws.append(field_names)  # 将模型字段名作为标题写入第一行
        for obj in queryset:  # 遍历选择的对象列表

            for field in field_names:
                data = [f'{getattr(obj, field)}' for field in field_names]  # 将模型属性值的文本格式组成列表
            ws.append(data)
        wb.save(response)
        return response

    export_as_excel.short_description = '导出Excel'  # 该动作在admin中的显示文字


admin.site.site_title = "高校大学生言论情感分析"
admin.site.site_header = "高校大学生言论情感分析"
admin.site.index_title = "高校大学生言论情感分析"
